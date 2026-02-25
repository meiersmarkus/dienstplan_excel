#!/usr/bin/env python3
"""Gruppenkalender VPA – trägt Dienste aller Personen für heute+morgen
in einen gemeinsamen CalDAV-Kalender ein.

Nutzt die gemeinsamen Module (config, downloader, laufzettel, event_builder,
holidays_de, utils) und enthält nur die Gruppen-spezifische Logik.
"""

import argparse
import datetime
import json
import logging
import os
import re
import signal
import sys
import time
from datetime import date, timedelta
from itertools import chain

import pytz
from caldav import DAVClient
from openpyxl import load_workbook

from config import AppConfig
from downloader import DownloadResult, download_plans
from event_builder import build_ical_event
from excel_parser import get_sorted_excel_files
from holidays_de import GermanHolidays
from laufzettel import LaufzettelManager
from utils import Timer, setup_logging, extract_date_from_filename

logger = logging.getLogger(__name__)

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TZ_BERLIN = pytz.timezone("Europe/Berlin")

# Kompilierte Regex-Patterns (einmal erstellt, überall wiederverwendet)
RE_DOT_TO_COLON = re.compile(r"(\b\d{2})\.(\d{2}\b)")
RE_TIME_RANGE = re.compile(r"^\d{2}:\d{2}\s*-\s*\d{2}:\d{2}\s*")
RE_CLEANUP = re.compile(r"\s*\(WT\)|\s*Info |\s+")
RE_TIME_SPACING = re.compile(r"(\b\d{2}:\d{2})\s*-\s*(\d{2}:\d{2}\b)")
RE_NAME_BRACKETS = re.compile(r"\s*[\r\n]*\(.*\)\s*[\r\n]*")
RE_IDENTIFIER = re.compile(r"^\d+\s*I\s*\d+$")


def parse_args():
    parser = argparse.ArgumentParser(description="Dienst zu Gruppenkalender VPA")
    parser.add_argument("-r", "--rewrite", action="store_true",
                        help="Alle vorhandenen Termine im Zeitbereich neu erstellen")
    parser.add_argument("-v", "--verbose", action="store_true",
                        help="Ausführliche Konsolenausgabe (DEBUG)")
    return parser.parse_args()


def load_schichten(folder_path: str) -> list:
    """Lädt die erlaubten Schichten aus vpa.json."""
    vpa_path = os.path.join(folder_path, "vpa.json")
    try:
        with open(vpa_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        schichten = data.get("schichten", [])
        # Flatten: Liste von Listen → flache Liste
        return [item for sublist in schichten for item in sublist]
    except Exception as e:
        logger.error("Fehler beim Laden von vpa.json: %s", e)
        return []


def connect_group_calendar(app_config: AppConfig) -> object:
    """Verbindet zum Gruppenkalender 'Dienstplan VPA'.

    Returns:
        caldav.Calendar-Objekt oder None bei Fehler.
    """
    service = "ard"
    try:
        creds = app_config.get_caldav_credentials(service)
    except ValueError as e:
        logger.error("Keine Credentials für Service '%s': %s", service, e)
        return None

    try:
        client = DAVClient(creds.base_url, username=creds.username, password=creds.password)
        principal = client.principal()
        calendar = principal.calendar(name="Dienstplan VPA")
        if calendar:
            return calendar
        logger.error("Kalender 'Dienstplan VPA' nicht gefunden.")
        return None
    except Exception as e:
        logger.error("CalDAV-Verbindungsfehler: %s", e)
        return None


def delete_old_events(calendar, days_back: int = 4):
    """Löscht alte Termine (vor gestern) aus dem Gruppenkalender."""
    heute = date.today()
    start_date = heute - timedelta(days=days_back)
    end_date = heute - timedelta(days=1)

    try:
        events = calendar.search(start=start_date, end=end_date, event=True)
        deleted = 0
        for event in events:
            try:
                event.delete()
                deleted += 1
            except Exception:
                pass
        if deleted:
            logger.debug("%d alte Termine gelöscht.", deleted)
    except Exception as e:
        logger.error("Fehler beim Löschen alter Termine: %s", e)


def match_workplace(shift_name: str, start_time: str, end_time: str,
                    shift_infos: list) -> str | None:
    """Findet den Arbeitsplatz aus dem Laufzettel für eine Schicht.

    Nutzt die ShiftInfo-Objekte aus dem LaufzettelManager.
    """
    cleaned = RE_CLEANUP.sub("", shift_name)

    for info in shift_infos:
        dienstname = (
            info.dienstname
            .replace("Samstag: ", "")
            .replace("Sonntag: ", "")
            .replace(" ", "")
            .strip()
        )

        if cleaned.lower() not in dienstname.lower():
            continue

        time_match = re.match(r"(\d{4})\s*-\s*(\d{4})", info.dienstzeit)
        if not time_match:
            continue

        html_start = f"{time_match.group(1)[:2]}:{time_match.group(1)[2:]}"
        html_end = f"{time_match.group(2)[:2]}:{time_match.group(2)[2:]}"

        if html_start == start_time and html_end == end_time:
            workplace = info.arbeitsplatz
            # Sonderbehandlungen
            if cleaned == "IngSchni" and workplace:
                workplace = workplace.split()[-1]
            if workplace == "Cut6 / Box2":
                workplace = "Cut6"
            return workplace

    return None


def process_timed_event(
    calendar, service_entry: str, work_date: date,
    name_without_brackets: str, laufzettel_mgr: LaufzettelManager,
    holidays: GermanHolidays, rewrite: bool,
):
    """Verarbeitet einen einzelnen zeitgebundenen Dienst für den Gruppenkalender.

    Returns:
        Log-Text bei neuem Eintrag, sonst None.
    """
    # Timeout-Schutz
    def timeout_handler(signum, frame):
        raise TimeoutError("Timeout bei Gruppenkalender-Verarbeitung.")

    signal.signal(signal.SIGALRM, timeout_handler)
    signal.alarm(300)

    try:
        time_match = re.match(r"(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})", service_entry)
        if not time_match:
            return None

        start_time_str = time_match.group(1)
        end_time_str = time_match.group(2)

        start_dt = datetime.datetime.strptime(
            f"{work_date.strftime('%Y-%m-%d')} {start_time_str}", "%Y-%m-%d %H:%M"
        )
        end_dt = datetime.datetime.strptime(
            f"{work_date.strftime('%Y-%m-%d')} {end_time_str}", "%Y-%m-%d %H:%M"
        )

        if start_dt.tzinfo is None:
            start_dt = TZ_BERLIN.localize(start_dt)
        if end_dt.tzinfo is None:
            end_dt = TZ_BERLIN.localize(end_dt)
        if end_dt < start_dt:
            end_dt += timedelta(days=1)

        # Titel zusammenbauen
        title = service_entry[time_match.end():].strip()
        title_cleaned = RE_CLEANUP.sub("", title)

        if title_cleaned == "Supervisor" and start_dt.hour == 9 and start_dt.minute == 30:
            full_title = f"{name_without_brackets}, {title} Büro"
        else:
            full_title = f"{name_without_brackets}, {title}"

        # Arbeitsplatz aus Laufzettel
        is_holiday, _ = holidays.is_holiday_or_weekend(work_date)
        werktags, wochenende = laufzettel_mgr.get_for_date(work_date)
        shift_infos = wochenende if is_holiday else werktags

        workplace = match_workplace(title, start_time_str, end_time_str, shift_infos)

        # Duplikat-Prüfung: nur Events dieser Person an diesem Tag
        existing_events = calendar.search(
            start=start_dt.replace(hour=0, minute=0, second=0),
            end=end_dt.replace(hour=23, minute=59, second=59),
            event=True,
        )
        event_exists = False

        for event in existing_events:
            try:
                event.load()
                event_summary = event.vobject_instance.vevent.summary.value
                event_start = event.vobject_instance.vevent.dtstart.value
                event_end = (
                    event.vobject_instance.vevent.dtend.value
                    if hasattr(event.vobject_instance.vevent, "dtend")
                    else None
                )
            except Exception:
                continue

            if not event_summary.startswith(name_without_brackets):
                continue

            # Timezone-Normalisierung
            if isinstance(event_start, datetime.date) and not isinstance(event_start, datetime.datetime):
                event_start = datetime.datetime.combine(event_start, datetime.time.min)
            if isinstance(event_start, datetime.datetime) and event_start.tzinfo is None:
                event_start = TZ_BERLIN.localize(event_start)
            if event_end and isinstance(event_end, datetime.date) and not isinstance(event_end, datetime.datetime):
                event_end = datetime.datetime.combine(event_end, datetime.time.min)
            if event_end and isinstance(event_end, datetime.datetime) and event_end.tzinfo is None:
                event_end = TZ_BERLIN.localize(event_end)

            # Rewrite-Modus: alles für diesen Tag löschen
            if rewrite and event_start.date() == start_dt.date():
                event.delete()
                continue

            # Exakter Match → Duplikat
            safe_title = full_title.replace("\n", " ").replace("\r", "").strip()
            safe_summary = event_summary.strip()
            if (safe_summary == safe_title
                    and event_start == start_dt
                    and event_end == end_dt):
                event_exists = True
                break

            # Anderer Termin derselben Person am selben Tag → löschen
            if event_start.date() == start_dt.date():
                logger.debug(
                    "Lösche '%s' am %s, weil ungleich '%s'.",
                    event_summary, start_dt.strftime("%d.%m.%Y"), full_title,
                )
                event.delete()

        if not event_exists:
            # Beschreibung bauen
            description = f"Dienst: {title_cleaned} von {name_without_brackets}, "
            if workplace:
                description += f"Platz: {workplace}, "
            else:
                description += "Platz: none, "
            description += "Alle Angaben und Inhalte sind ohne Gewähr. "
            description += f"Änderungsdatum: {datetime.datetime.now().strftime('%d.%m.%Y, %H:%M')}"

            ical_data = build_ical_event(
                title=full_title,
                start=start_dt,
                end=end_dt,
                description=description,
            )
            calendar.add_event(ical_data)
            log_text = (
                f"{start_dt.strftime('%d.%m.%Y')}, "
                f"{start_dt.strftime('%H:%M')} bis {end_dt.strftime('%H:%M')}: "
                f"{full_title}, {workplace}"
            )
            logger.debug("[Dienst] %s", log_text)
            return log_text

        return None

    finally:
        signal.alarm(0)


def process_excel_file(
    file_path: str, heute: date, schichten: list,
    calendar, laufzettel_mgr: LaufzettelManager,
    holidays: GermanHolidays, rewrite: bool,
) -> list:
    """Verarbeitet eine Excel-Datei für den Gruppenkalender.

    Iteriert spaltenweise (pro Tag) und nimmt alle Personen mit.
    Nur heute und morgen werden verarbeitet.

    Returns:
        Liste der neu eingetragenen Termine als Log-Texte.
    """
    new_entries = []

    try:
        wb = load_workbook(file_path, data_only=True)
        ws = wb.active
    except Exception as e:
        logger.error("Fehler beim Laden von %s: %s", os.path.basename(file_path), e)
        return new_entries

    # Identifikationszeile finden (enthält "I" für Kalenderwochen)
    identifier_row_index = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val and isinstance(str(val), str):
            cleaned = str(val).strip()
            if RE_IDENTIFIER.match(cleaned):
                identifier_row_index = r
                break

    if identifier_row_index is None:
        wb.close()
        return new_entries

    morgen = heute + timedelta(days=1)

    for day_col in range(2, 9):  # Spalten B (2) bis H (8)
        cell_date = ws.cell(row=identifier_row_index, column=day_col).value

        if isinstance(cell_date, datetime.datetime):
            work_date = cell_date.date()
        elif isinstance(cell_date, datetime.date):
            work_date = cell_date
        else:
            continue

        # Nur heute und morgen
        if work_date != heute and work_date != morgen:
            continue

        for r in range(identifier_row_index + 1, ws.max_row + 1):
            cell_val = ws.cell(row=r, column=day_col).value
            if cell_val is None:
                continue

            raw_entry = str(cell_val)
            service_entry = RE_DOT_TO_COLON.sub(r"\1:\2", raw_entry)
            schicht_key = RE_CLEANUP.sub("", RE_TIME_RANGE.sub("", service_entry))

            if schicht_key not in schichten:
                if "Projekt" not in service_entry and "Bereitschaft" not in service_entry:
                    continue

            if "Projekt" in service_entry or "Bereitschaft" in service_entry:
                if r >= 86:
                    continue
                service_entry = f"09:00 - 09:01 {service_entry}"

            service_entry = service_entry.replace("\n", " ").replace("\r", " ")
            service_entry = re.sub(r" {2,}", " ", service_entry)
            service_entry = RE_TIME_SPACING.sub(r"\1 - \2", service_entry)

            # Name aus Spalte A
            name_val = ws.cell(row=r, column=1).value
            name = str(name_val) if name_val else ""
            name_without_brackets = RE_NAME_BRACKETS.sub("", name).strip()

            if not name_without_brackets:
                continue

            if RE_TIME_SPACING.search(service_entry):
                log_text = process_timed_event(
                    calendar, service_entry, work_date,
                    name_without_brackets, laufzettel_mgr,
                    holidays, rewrite,
                )
                if log_text:
                    new_entries.append(log_text)

    wb.close()
    return new_entries


def main():
    args = parse_args()
    console_level = logging.DEBUG if args.verbose else logging.INFO
    setup_logging(BASE_DIR, console_level=console_level)

    logger.info("Starte Gruppenkalenderaktualisierung VPA...")

    with Timer("Gesamtdauer", log_threshold_seconds=0):
        # Feiertags-Check
        feiertage = GermanHolidays()
        is_free, holiday_name = feiertage.is_holiday_or_weekend(date.today())
        if is_free and time.localtime().tm_hour > 2:
            logger.info(
                "%s ist %s. Skript wird nicht ausgeführt.",
                date.today().strftime("%d.%m.%Y"), holiday_name,
            )
            return

        # Konfiguration laden
        app_config = AppConfig(BASE_DIR)

        # Download über das gemeinsame Modul
        with Timer("Download"):
            result = download_plans(app_config, BASE_DIR, fast=True)
        if result == DownloadResult.CONNECTION_ERROR:
            logger.error("Server nicht erreichbar. Abbruch.")
            sys.exit(2)

        # CalDAV-Verbindung zum Gruppenkalender
        with Timer("CalDAV-Verbindung", log_threshold_seconds=5):
            calendar = connect_group_calendar(app_config)
        if not calendar:
            logger.error("Konnte nicht zum Gruppenkalender verbinden. Abbruch.")
            sys.exit(1)

        # Alte Termine löschen
        delete_old_events(calendar)

        # Laufzettel und Feiertage
        laufzettel_mgr = LaufzettelManager(BASE_DIR)

        # Schichten-Filter laden
        schichten = load_schichten(BASE_DIR)

        # Excel-Dateien verarbeiten
        plans_folder = os.path.join(BASE_DIR, "Plaene", "MAZ_TAZ Dienstplan")
        xlsx_files = get_sorted_excel_files(plans_folder)

        if not xlsx_files:
            logger.debug("Keine .xlsx-Dateien gefunden.")
            return

        heute = date.today()
        all_new_entries = []

        for file_path in xlsx_files:
            new_entries = process_excel_file(
                file_path, heute, schichten,
                calendar, laufzettel_mgr, feiertage, args.rewrite,
            )
            all_new_entries.extend(new_entries)

        if all_new_entries:
            logger.info("%d neue Termine eingetragen.", len(all_new_entries))


if __name__ == "__main__":
    main()
