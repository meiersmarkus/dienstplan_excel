"""Excel-Parser: Liest Dienstpläne und extrahiert Dienst-Einträge pro Benutzer."""

import datetime
import logging
import os
import re
from dataclasses import dataclass
from itertools import chain
from typing import List, Optional, Tuple

from openpyxl import load_workbook

from utils import to_datetime, extract_date_from_filename

logger = logging.getLogger(__name__)

# openpyxl-Warnungen zu Zeichnungen unterdrücken
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.reader.drawings")


@dataclass
class ShiftEntry:
    """Ein einzelner Diensteintrag aus dem Excel."""
    date: datetime.datetime
    raw_text: str           # Originaltext aus der Zelle
    start_time: Optional[str] = None  # "HH:MM" oder None bei ganztägig
    end_time: Optional[str] = None    # "HH:MM" oder None bei ganztägig
    shift_name: Optional[str] = None  # z.B. "OMSchni 3", "Supervisor"
    is_timed: bool = False  # True wenn Uhrzeiten vorhanden

    @property
    def is_all_day(self) -> bool:
        return not self.is_timed


def parse_excel_file(file_path: str, user_name: str) -> Tuple[List[ShiftEntry], bool]:
    """Parst eine einzelne Excel-Datei und extrahiert Dienste für einen Benutzer.

    Args:
        file_path: Pfad zur .xlsx-Datei
        user_name: Name des Benutzers (wie in colleagues.json, z.B. "Meier, M.")

    Returns:
        (liste_von_shift_entries, user_found)
        user_found=False bedeutet: Benutzer nicht im Plan → ggf. FT/Löschen
    """
    try:
        wb = load_workbook(file_path, data_only=True)
    except Exception as e:
        logger.error("Kann %s nicht öffnen: %s", os.path.basename(file_path), e)
        return [], False

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # 1. Identifikationszeile finden (enthält Wochen-Nummern wie "40  I  41")
    identifier_row = _find_identifier_row(rows)
    if identifier_row is None:
        logger.error(
            "Keine Datumszeile in %s gefunden.", os.path.basename(file_path)
        )
        return [], False

    # 2. Benutzer-Zeile suchen
    user_row = _find_user_row(rows, user_name)

    # 3. Datumsspalten extrahieren (Index 1-7 = Mo-So)
    dates = _extract_week_dates(identifier_row)

    if user_row is None:
        # Benutzer nicht gefunden → Daten für die Woche zurückgeben mit FT
        entries = []
        for date_val in dates:
            if date_val:
                entries.append(ShiftEntry(date=date_val, raw_text="FT"))
        return entries, False

    # 4. Dienste der Benutzer-Zeile extrahieren
    entries = []
    for i, date_val in enumerate(dates):
        if date_val is None:
            continue

        cell_value = user_row[i + 1]  # +1 weil Index 0 = Name-Spalte
        raw_text = _clean_cell_value(cell_value)

        entry = _parse_shift_entry(raw_text, date_val)
        entries.append(entry)

    return entries, True


def get_sorted_excel_files(folder_path: str) -> List[str]:
    """Findet alle .xlsx-Dateien im Ordner und sortiert sie nach Datum im Dateinamen.

    Dateien mit erkennbarem Datum kommen zuerst (chronologisch),
    dann Dateien ohne erkennbares Datum.
    """
    xlsx_files = []
    for root, _, files in os.walk(folder_path):
        for f in files:
            if f.endswith(".xlsx") and "test" not in f.lower():
                xlsx_files.append(os.path.join(root, f))

    with_date = []
    without_date = []
    for fpath in xlsx_files:
        file_date = extract_date_from_filename(os.path.basename(fpath))
        if file_date:
            with_date.append((fpath, file_date))
        else:
            without_date.append(fpath)

    with_date.sort(key=lambda x: x[1])
    return [f for f, _ in with_date] + without_date


# ---------------------------------------------------------------------------
# Interne Hilfsfunktionen
# ---------------------------------------------------------------------------

def _find_identifier_row(rows) -> Optional[tuple]:
    """Findet die Zeile mit den Kalenderwochen (z.B. '40  I  41')."""
    pattern = re.compile(r"^\d+\s*I\s*\d+$")
    for row in rows:
        if row[0] and isinstance(row[0], str):
            cleaned = row[0].strip()
            if pattern.match(cleaned):
                return row
    return None


def _find_user_row(rows, user_name: str) -> Optional[tuple]:
    """Sucht die Zeile eines Benutzers in der Excel-Tabelle."""
    target_full = _clean_excel_name(user_name)
    # Kurzform: Nachname ohne ", X." am Ende
    target_short = _clean_excel_name(re.sub(r",\s*[A-Z]\.?$", "", user_name).strip())

    for row in rows:
        cell_val = row[0]
        if not cell_val:
            continue

        row_name = _clean_excel_name(cell_val)
        if not row_name:
            continue

        if row_name == target_full or row_name == target_short:
            return row

    return None


def _clean_excel_name(name) -> str:
    """Bereinigt einen Namen aus der Excel-Tabelle für den Vergleich."""
    if not name:
        return ""
    try:
        if isinstance(name, bytes):
            name = name.decode("utf-8", errors="ignore")
        name = str(name)
    except Exception:
        return ""

    # Inhalte in Klammern entfernen (z.B. "(TV)", "(fester Freier)")
    name = re.sub(r"\s*\(.*?\)", "", name)
    # Unicode-Sonderzeichen normalisieren
    name = name.replace("\u00A0", " ").replace("\u200b", "").replace("\ufeff", "")
    # Mehrfache Leerzeichen reduzieren
    name = re.sub(r"\s+", " ", name)
    return name.strip().lower()


def _extract_week_dates(identifier_row) -> List[Optional[datetime.datetime]]:
    """Extrahiert die 7 Datumswerte aus der Identifikationszeile (Index 1-7)."""
    dates = []
    for i in range(1, 8):
        val = identifier_row[i] if i < len(identifier_row) else None
        if val is not None:
            dates.append(to_datetime(val))
        else:
            dates.append(None)
    return dates


def _clean_cell_value(value) -> str:
    """Bereinigt den Wert einer Dienst-Zelle."""
    if value is None:
        return "FT"
    if isinstance(value, float):
        return "FT"  # NaN-Werte
    if not isinstance(value, str):
        return "FT"

    text = value.replace("\n", " ").replace("\r", " ")
    text = re.sub(r"\s+", " ", text)  # Doppel-Leerzeichen
    # Formate wie 10.00 zu 10:00 reparieren
    text = re.sub(r"(\b\d{2})\.(\d{2}\b)", r"\1:\2", text)
    # Leerzeichen um Bindestrich normalisieren
    text = re.sub(r"(\b\d{2}:\d{2})\s*-\s*(\d{2}:\d{2}\b)", r"\1 - \2", text)
    return text.strip()


_TIME_PATTERN = re.compile(r"(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})")


def _parse_shift_entry(raw_text: str, date_val: datetime.datetime) -> ShiftEntry:
    """Parst einen bereinigten Zelleninhalt in einen ShiftEntry."""
    match = _TIME_PATTERN.search(raw_text)
    if match:
        start_time = match.group(1)
        end_time = match.group(2)
        # Schichtname steht nach der Zeitangabe
        shift_name = raw_text[match.end():].strip() or "Kein Titel"
        return ShiftEntry(
            date=date_val,
            raw_text=raw_text,
            start_time=start_time,
            end_time=end_time,
            shift_name=shift_name,
            is_timed=True,
        )
    else:
        return ShiftEntry(
            date=date_val,
            raw_text=raw_text,
            shift_name=raw_text.strip() or "FT",
        )
