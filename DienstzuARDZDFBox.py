import sys
import time
import os
import argparse
import json
from openpyxl import load_workbook
import datetime
import locale
import re
from bs4 import BeautifulSoup
from caldav import DAVClient
from datetime import date, timedelta
import datetime as dt
import pytz
import holidays
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from dateutil.easter import easter
from itertools import chain
import logging
import urllib.parse
import warnings
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl.reader.drawings")

if sys.platform.startswith('win'):
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')
    except AttributeError:
        pass # Ältere Python-Versionen ignorieren

class CalendarCache:
    def __init__(self, client, calendar_obj):
        self.client = client
        self.calendar = calendar_obj
        self.events_by_date = {}  # Key: datetime.date, Value: List of events
        self.all_events_flat = [] # Flache Liste für schnelle Iteration

    def load_all_events(self, start_date, end_date):
        # print(f"[INFO] Lade Kalenderdaten vom {start_date.strftime('%d.%m.%Y')} bis {end_date.strftime('%d.%m.%Y')}...")
        try:
            # expand=False ist wichtig für Performance!
            events = self.calendar.search(start=start_date, end=end_date, event=True, expand=False)
            
            self.events_by_date = {}
            self.all_events_flat = []
            
            for event in events:
                self._add_to_local_cache(event)
                
            # print(f"[INFO] {len(self.all_events_flat)} Termine im Speicher.")
        except Exception as e:
            print(f"[ERROR] Fehler beim Laden des Kalender-Caches: {e}")

    def _add_to_local_cache(self, event):
        """Hilfsfunktion: Fügt ein Event-Objekt in die internen Strukturen ein."""
        try:
            if not hasattr(event, 'vobject_instance') or event.vobject_instance is None:
                if not event.data:
                    return            
            start = event.vobject_instance.vevent.dtstart.value
            if isinstance(start, datetime.datetime):
                date_key = start.date()
            else:
                date_key = start
            
            if date_key not in self.events_by_date:
                self.events_by_date[date_key] = []

            self.events_by_date[date_key].append(event)
            self.all_events_flat.append(event)
        except Exception as e:
            # Abfangen von korrupten Events
            print(f"[DEBUG] Cache Fehler bei Event: {e}")
            pass

    def get_events_on_date(self, check_date):
        """Gibt eine Liste von Events für ein bestimmtes Datum zurück."""
        if isinstance(check_date, datetime.datetime):
            check_date = check_date.date()
        return self.events_by_date.get(check_date, [])

    def add_event(self, ical_data):
        """Fügt Event zum Server UND zum lokalen Cache hinzu."""
        try:
            new_event = self.calendar.add_event(ical_data)
            self._add_to_local_cache(new_event)
            return new_event
        except Exception as e:
            print(f"[ERROR] Fehler beim Hinzufügen zum Kalender: {e}")
            return None

    def delete_event(self, event):
        """Löscht Event vom Server UND aus dem lokalen Cache."""
        try:
            # 1. Aus Cache entfernen (bevor wir es auf dem Server löschen und Daten verlieren)
            try:
                start = event.vobject_instance.vevent.dtstart.value
                if isinstance(start, datetime.datetime):
                    date_key = start.date()
                else:
                    date_key = start
                
                if date_key in self.events_by_date:
                    if event in self.events_by_date[date_key]:
                        self.events_by_date[date_key].remove(event)
            except:
                pass # Falls Zugriff auf vobject fehlschlägt

            if event in self.all_events_flat:
                self.all_events_flat.remove(event)
            event.delete()
        except Exception as e:
            print(f"[ERROR] Fehler beim Löschen des Events: {e}")

# Globale Variable
cal_cache = None

# Initialisiere Timer
timers = {}
logging.getLogger("caldav").disabled = True


def start_timer(timer_name):
    timers[timer_name] = time.time()


start_timer("gesamt")
start_timer("initial")


def end_timer(timer_name, task_description):
    if timer_name in timers:
        end_time = time.time()
        elapsed_time = end_time - timers[timer_name]
        if not (timer_name in ("caldav", "initial", "gesamt") and elapsed_time <= 20):
            print(f"[TIME] {task_description}: {elapsed_time:.2f} Sekunden")
        del timers[timer_name]  # Timer entfernen, wenn er fertig ist
    else:
        print(f"[ERROR] Kein aktiver Timer mit dem Namen: {timer_name}")

def encode_calendar_url(calendar_name):
    return urllib.parse.quote(calendar_name)

def replace_german_umlauts(text):
    """Ersetzt deutsche Umlaute und Sonderzeichen für URLs."""
    replacements = {
        'ä': '', 'ö': '', 'ü': '', 'ß': '',
        'Ä': '', 'Ö': '', 'Ü': ''
    }
    for k, v in replacements.items():
        text = text.replace(k, v)
    return text


def is_holiday_or_weekend(datum):
    if datum in de_holidays:
        return True, de_holidays.get(datum)
    if datum.weekday() >= 5:
        return True, "Samstag" if datum.weekday() == 5 else "Sonntag"
    return False, None


def parse_html_for_workplace_info_with_cache(html_file_path):
    global laufzettel_cache
    if html_file_path in laufzettel_cache:
        return laufzettel_cache[html_file_path]
    
    # Wenn nicht im Cache, parse die Datei und speichere das Ergebnis
    laufzettel_werktags, laufzettel_we = parse_html_for_workplace_info(html_file_path)
    laufzettel_cache[html_file_path] = (laufzettel_werktags, laufzettel_we)
    # print(f"[DEBUG] Laufzettel-Cache aktualisiert für {html_file_path}")
    return laufzettel_werktags, laufzettel_we


def parse_html_for_workplace_info(html_file_path):  # Function to parse HTML and extract workplace, breaks, and tasks
    # start_timer("html")
    def extract_info(table):
        # Extrahiert die Informationen aus einer HTML-Tabelle und gibt eine Liste von Einträgen zurück.
        info = []
        if table:
            for row in table.find_all('tr'):
                cols = row.find_all('td')
                if len(cols) >= 5:
                    info.append({
                        'dienstname': cols[0].text.replace('\u00A0', ' ').strip(),
                        'dienstzeit': cols[1].text.replace('\u00A0', ' ').strip(),
                        'arbeitsplatz': cols[2].text.replace('\u00A0', ' ').strip(),
                        'pausenzeit': cols[3].text.replace('\u00A0', ' ').strip(),
                        'task': cols[4].text.replace('\u00A0', ' ').strip()
                    })
                    if len(info) > 40:
                        break
        for entry in info:
            if re.search(r"[\u00A0\u200B\uFEFF]", str(entry)):
                print("[DEBUG] Unsichtbare Zeichen gefunden:", repr(entry))
        return info

    # HTML parsen
    with open(html_file_path, 'r', encoding='utf-8') as file:
        soup = BeautifulSoup(file, 'html.parser')

    # Werktags- und Wochenend-Bereich extrahieren
    werktags_table = soup.find('table', summary='Laufzettel ab 01.01.23 Werktags')
    wochenende_table = soup.find('table', summary='Laufzettel ab 01.01.23 Wochenende')

    laufzettel_werktags = extract_info(werktags_table) if werktags_table else []
    laufzettel_we = extract_info(wochenende_table) if wochenende_table else []

    # Rückgabe der getrennten Listen
    # end_timer("html", "HTML auslesen")
    return laufzettel_werktags, laufzettel_we


def to_python_datetime(value):
    """Ersetzt pd.to_datetime für Einzelwerte."""
    if value is None:
        return None
    if isinstance(value, (datetime.datetime, datetime.date)):
        # Wenn es schon ein Datum ist (Excel liefert das oft direkt)
        if isinstance(value, datetime.date) and not isinstance(value, datetime.datetime):
            return datetime.datetime.combine(value, datetime.time.min)
        return value
    if isinstance(value, str):
        # Versuche gängige Formate
        for fmt in ('%Y-%m-%d', '%d.%m.%Y', '%Y%m%d'):
            try:
                return datetime.datetime.strptime(value, fmt)
            except ValueError:
                continue
    raise ValueError(f"Konnte Datum nicht parsen: {value}")


def count_night_shifts(search_until_date):
    # search_until_date: Bis zu diesem Datum zählen (exklusive oder inklusive, je nach Logik)
    # Wir iterieren über die flache Liste im Cache
    
    if isinstance(search_until_date, str):
        search_until_date = to_python_datetime(search_until_date)
    
    # Sicherstellen, dass wir nur bis zum Tag vor dem aktuellen Eintrag schauen 
    # (oder inklusive, je nachdem wie deine alte Logik war. Hier: bis Mittag des Tages)
    limit_date = search_until_date.date()
    current_year = limit_date.year

    count = 0
    for event in cal_cache.all_events_flat:
        try:
            start = event.vobject_instance.vevent.dtstart.value
            if isinstance(start, datetime.datetime):
                s_date = start.date()
                s_time = start.time()
            else:
                s_date = start
                s_time = datetime.time(0, 0)

            # Filter: Gleiches Jahr UND Datum <= aktuelles Datum
            if s_date.year == current_year and s_date < limit_date:
                # Nachtschicht Logik (> 20:00 Uhr)
                if s_time >= datetime.time(20, 0):
                    count += 1
        except:
            continue
            
    return count


def create_ical_event(
    title,
    start_datetime,
    end_datetime=None,
    all_day=False,
    description=None
):
    if any(term in title for term in ['FT']) and user_name == load_credentials("user2", config_path):
        # print(f"[DEBUG] {title} übersprungen, keine FT ausgewählt wurde.")
        return
    if any(term in title for term in ['FT', 'UR', 'NV', 'KD', 'KR']) and dienste:
        # print(f"[DEBUG] {title} übersprungen, weil -d ausgewählt wurde.")
        return
    try:
        now = datetime.datetime.now(pytz.timezone("Europe/Berlin"))
        location = ''
        if all_day:
            # For all-day events, only include the date, without time and timezone
            start_str = start_datetime.strftime('%Y%m%d')
            end_str = (start_datetime + datetime.timedelta(days=1)).strftime('%Y%m%d')
            dtstart_str = f"DTSTART;VALUE=DATE:{start_str}"
            dtend_str = f"DTEND;VALUE=DATE:{end_str}"
        else:
            # Normal events with time and timezone
            start_str = start_datetime.strftime('%Y%m%dT%H%M%S')
            end_str = end_datetime.strftime('%Y%m%dT%H%M%S') if end_datetime else start_str
            dtstart_str = f"DTSTART;TZID=Europe/Berlin:{start_str}"
            dtend_str = f"DTEND;TZID=Europe/Berlin:{end_str}"
            if ort:
                location = r'LOCATION:Hugh-Greene-Weg 1\, 22529 Hamburg'

        # Use the provided description or default to Dienst information
        description_str = (
            description if description else
            f"Eintrag: {title}, Alle Angaben und Inhalte sind ohne Gewähr. "
            f"Änderungsdatum: {datetime.datetime.now().strftime('%d.%m.%Y, %H:%M')}"
        )
        busy = (
            "X-MICROSOFT-CDO-BUSYSTATUS:OOF"
            if any(term in title for term in ["FT", "UR", "NV", "KD", "KR", "FU", "AS"])
            else "X-MICROSOFT-CDO-BUSYSTATUS:BUSY"
        )
        transparent = "TRANSP:TRANSPARENT" if any(
            term in title for term in ["FT", "UR", "NV", "KD", "KR", "FU", "AS"]
        ) else "TRANSP:OPAQUE"
        sanitized_title = title.replace("\n", " ").replace("\r", "").strip()
        sanitized_desc = description_str.replace("\n", " ").replace("\r", "").strip()
        ical_event = f"""BEGIN:VCALENDAR
CALSCALE:GREGORIAN
VERSION:2.0
PRODID:-//meiersmarkus//NONSGML v1.0//DE
BEGIN:VEVENT
SUMMARY:{sanitized_title}
{transparent}
{dtstart_str}
{dtend_str}
DTSTAMP:{now.strftime('%Y%m%dT%H%M%SZ')}
UID:{now.timestamp()}@meiersmarkus.de
SEQUENCE:1
DESCRIPTION:{sanitized_desc}
LAST-MODIFIED:{now.strftime('%Y%m%dT%H%M%SZ')}
{location}
{busy}
END:VEVENT
BEGIN:VTIMEZONE
TZID:Europe/Berlin
BEGIN:DAYLIGHT
TZOFFSETFROM:+0100
TZOFFSETTO:+0200
TZNAME:CEST
DTSTART:19700329T020000
RRULE:FREQ=YEARLY;BYMONTH=3;BYDAY=-1SU
END:DAYLIGHT
BEGIN:STANDARD
TZOFFSETFROM:+0200
TZOFFSETTO:+0100
TZNAME:CET
DTSTART:19701025T030000
RRULE:FREQ=YEARLY;BYMONTH=10;BYDAY=-1SU
END:STANDARD
END:VTIMEZONE
END:VCALENDAR
"""
        # Termine in die Liste einfügen
        zeit = f" von {start_datetime.strftime('%H:%M')} - {end_datetime.strftime('%H:%M')} Uhr" if not all_day else ""
        eingetragene_termine.append(f"{start_datetime.strftime('%d.%m.%Y')}{zeit}: {title}")

        return ical_event
    except Exception as e:
        print(f"[ERROR] Fehler beim Erstellen des Events: {e}")
        return None


def process_all_day_event(service_entry, start_date):  # Funktion zur Verarbeitung eines ganztägigen Events
    title = service_entry.strip() or "Ganztägiger Termin"
    start_datetime = to_python_datetime(start_date)

    try:
        # Suche nach vorhandenen ganztägigen Terminen
        existing_events = cal_cache.get_events_on_date(start_datetime.date())
        event_exists = False

        for event in existing_events[:]:
            event_summary = event.vobject_instance.vevent.summary.value
            event_start = event.vobject_instance.vevent.dtstart.value
            event_end = (event.vobject_instance.vevent.dtend.value
                         if hasattr(event.vobject_instance.vevent, 'dtend')
                         else None)

            if any(term in event_summary for term in ['FT']) and user_name == load_credentials("user2", config_path):
                print(f"[DEBUG] Lösche {event_summary} vom {event_start}, da FT nicht eingetragen werden sollen.")
                cal_cache.delete_event(event)
                continue
            if any(term in event_summary for term in ['FT', 'UR', 'NV', 'KD', 'KR']) and dienste:
                print(f"[DEBUG] Lösche {event_summary} vom {event_start}, da nur Dienste eingetragen werden sollen.")
                cal_cache.delete_event(event)
                continue

            # Prüfen, ob event_start eine Uhrzeit enthält
            if isinstance(event_start, dt.datetime):
                event_start = event_start.date()  # Nur das Datum extrahieren

            if rewrite and event_start == start_datetime.date():
                cal_cache.delete_event(event)
                continue
            elif event_summary == title.replace("\n", " ").replace("\r", "").strip() and event_start == start_datetime.date():
                event_exists = True
                break
            elif event_start == start_datetime.date():
                if isinstance(event_start, datetime.datetime) and isinstance(event_end, datetime.datetime):
                    print(f"[DEBUG] {start_datetime.strftime('%d.%m.%Y')}, {event_start.strftime('%H:%M')} bis {event_end.strftime('%H:%M')} '{event_summary}' wird gelöscht.")
                else:
                    print(f"[DEBUG] {start_datetime.strftime('%d.%m.%Y')}, '{event_summary}' wird gelöscht.")
                cal_cache.delete_event(event)

        # Event erstellen, wenn kein passender Termin vorhanden ist
        if not event_exists:
            ical_data = create_ical_event(title, start_datetime, all_day=True)
            if ical_data:
                cal_cache.add_event(ical_data)
                print(f"[Dienst] {start_datetime.strftime('%d.%m.%Y')}: {title}")

    except Exception as e:
        print(f"[ERROR] Fehler beim Speichern oder Löschen des ganztägigen Events: {title} am {start_datetime}")
        print(e)

def normalize_string(s):
    return re.sub(r'\s+', ' ', s.strip())

# Funktion zur Verarbeitung eines zeitgebundenen Events
def process_timed_event(service_entry, start_date, laufzettel_werktags, laufzettel_we, countnightshifts, nonightshifts):
    # Extract start and end time from Excel entry
    time_match = re.match(r'(\d{2}:\d{2})\s*-\s*(\d{2}:\d{2})', service_entry)
    if time_match:
        start_time_str = time_match.group(1)
        end_time_str = time_match.group(2)

        start_datetime = datetime.datetime.strptime(f"{start_date.strftime('%Y-%m-%d')} {start_time_str}", '%Y-%m-%d %H:%M')
        end_datetime = datetime.datetime.strptime(f"{start_date.strftime('%Y-%m-%d')} {end_time_str}", '%Y-%m-%d %H:%M')

        if user_name == load_credentials("user1", config_path) and end_datetime.time() < datetime.time(8, 0):
            end_datetime = tz_berlin.localize(
                datetime.datetime.combine(end_datetime.date(), datetime.time(23, 59)), is_dst=None
            )
        elif end_datetime < start_datetime:
            end_datetime += datetime.timedelta(days=1)

        if start_datetime.tzinfo is None:
            start_datetime = tz_berlin.localize(start_datetime)
        if end_datetime.tzinfo is None:
            end_datetime = tz_berlin.localize(end_datetime, is_dst=None)          

        # Get the basic title from the Excel entry (e.g., "Schnitt 2")
        title = service_entry[time_match.end():].strip() or "Kein Titel"
        # print(f"[DEBUG] Excel event: {title}, start: {start_time_str}, end: {end_time_str}")

        # Default values for workplace info (in case no match is found in HTML)
        workplace = None
        break_time = None
        task = None

        is_holiday_flag, holiday_name = is_holiday_or_weekend(start_date.date())
        workplace_info = laufzettel_we if is_holiday_flag else laufzettel_werktags
        # print(f"[DEBUG] Verwende {'Wochenende' if is_holiday_flag else 'Werktags'}-Tabelle für {start_date.strftime('%d.%m.%Y')}")
        try:
            # Entferne "(WT)", "Info" und alle Leerzeichen aus dem zu vergleichenden String
            cleaned_service_entry = re.sub(r'\s*\(WT\)|\s*Info |\s+', '', service_entry[time_match.end():].strip())
            for info in workplace_info:
                dienstname = info['dienstname'].replace("Samstag: ", "").replace("Sonntag: ", "").replace(" ", "").strip()
                # print(f"[DEBUG] Vergleiche Excel '{cleaned_service_entry}' mit Laufzettel '{dienstname}'")
                if cleaned_service_entry.lower() in dienstname.lower().strip():
                    # print(f"[DEBUG] {dienstname} gefunden. Dienstzeit: {info['dienstzeit']}")
                    # DIENSTZEIT ist im HHMM-HHMM Format
                    # print(f"[DEBUG] Dienstzeit: {info['dienstzeit']}")
                    html_time_match = re.match(r'(\d{4})\s*-\s*(\d{4})', info['dienstzeit'])
                    if html_time_match:
                        html_start_time = f"{html_time_match.group(1)[:2]}:{html_time_match.group(1)[2:]}"
                        html_end_time = f"{html_time_match.group(2)[:2]}:{html_time_match.group(2)[2:]}"
                        # print(f"[DEBUG] HTML Startzeit: {html_start_time}, HTML Endzeit: {html_end_time}")

                        if html_start_time == start_time_str and html_end_time == end_time_str:
                            workplace = info.get('arbeitsplatz', None)
                            break_time = info.get('pausenzeit', None)
                            task = info.get('task', None)
                            # print(f"[DEBUG] {cleaned_service_entry}, {workplace}")
                            break
            # print(f"[DEBUG] Am {start_date.strftime('%d.%m.%Y')}: {title}, Workplace: {workplace}, Break: {break_time}, Task: {task}")

            if user_name == load_credentials("user1", config_path):
                full_title = f"{start_time_str}-{end_time_str} {title}"
            else:
                full_title = f"{title}, {workplace}" if workplace and workplace not in title else title

            if start_datetime.time() >= datetime.time(20, 0) and nonightshifts == False:  # Dienste mit Startzeit nach 20:00 Uhr
                # Prüfen, ob die Nachtschichtzählung neu gestartet werden soll
                if not countnightshifts or start_date.month in [1, 2, 3, 4]:
                    # print(f"[DEBUG] Starte neue Zählung: Monat={start_date.month}, Datum={start_date}")
                    night_shifts_count = count_night_shifts(start_date)
                    night_shifts_count += 1  # Erhöhe die Anzahl um 1 für den aktuellen Dienst
                    countnightshifts = True  # Markiere, dass die Zählung gestartet wurde
                else:
                    night_shifts_count += 1  # Erhöhe die Anzahl um 1 für den aktuellen Dienst
                    # print(f"[DEBUG] Nachtschicht-Zähler: {night_shifts_count}")
                full_title += f" ({night_shifts_count})"

            # print(f"[DEBUG] Full event title: {full_title}")
            # print(f"[DEBUG] Excel: '{full_title.strip()}' am '{start_datetime.date()}' von '{start_datetime.time()}' bis '{end_datetime.time()}' Uhr.")
            # print(f"[DEBUG] Prüfe '{full_title.strip()}' am '{start_datetime.date()}' mit Laufzetteldatei vom '{current_laufzettel.strftime('%d.%m.%Y')}'")  
            # Now check if the event with the full title already exists
            existing_events = cal_cache.get_events_on_date(start_datetime.date())

            event_exists = False

            for event in existing_events[:]:
                event_summary = event.vobject_instance.vevent.summary.value
                event_start = event.vobject_instance.vevent.dtstart.value
                event_end = (event.vobject_instance.vevent.dtend.value
                             if hasattr(event.vobject_instance.vevent, 'dtend')
                             else None)
                # print(f"[DEBUG] {len(existing_events)} Termine gefunden.")

                if any(term in event_summary for term in ['FT']) and user_name == load_credentials("user2", config_path):
                    print(f"[DEBUG] Lösche {event_summary} vom {event_start}, da FT nicht eingetragen werden sollen.")
                    cal_cache.delete_event(event)
                    continue
                if any(term in event_summary for term in ['FT', 'UR', 'NV', 'KD', 'KR']) and dienste:
                    print(f"[DEBUG] Lösche {event_summary} vom {event_start}, da nur Dienste eingetragen werden sollen.")
                    cal_cache.delete_event(event)
                    continue

                # Ensure event_start and event_end are datetime objects, and localize if necessary
                if isinstance(event_start, datetime.date) and not isinstance(event_start, datetime.datetime):
                    event_start = datetime.datetime.combine(event_start, datetime.time.min)
                if isinstance(event_start, datetime.datetime) and event_start.tzinfo is None:
                    event_start = tz_berlin.localize(event_start)
                if event_end and isinstance(event_end, datetime.date) and not isinstance(event_end, datetime.datetime):
                    event_end = datetime.datetime.combine(event_end, datetime.time.min)
                if event_end and isinstance(event_end, datetime.datetime) and event_end.tzinfo is None:
                    event_end = tz_berlin.localize(event_end)

                if rewrite:
                    if event_start.date() == start_datetime.date():
                        cal_cache.delete_event(event)
                        continue
                # Compare the fully generated title with the existing event's summary
                # print(f"[DEBUG] Vergleiche Kalender: '{event_summary}' am '{event_start}' bis {event_end}  mit Excel: '{full_title}' am '{start_datetime}' bis '{end_datetime}'.")
                if (event_summary.replace("\n", " ").replace("\r", "").strip() == full_title.replace("\n", " ").replace("\r", "").strip() and
                        event_start == start_datetime and
                        event_end == end_datetime):
                    # print(f"[DEBUG] Event '{full_title}' am {start_datetime.strftime('%d.%m.%Y')} bereits vorhanden.")
                    event_exists = True
                    # print(f"[DEBUG] Event '{full_title}' already exists. Skipping creation.")
                    break
                # print(f"[DEBUG] Excel: '{full_title.strip()}' am '{start_datetime.date()}'. "f"Kalender: '{event_summary}' am '{event_start.date()}'.")
                if event_start.date() == start_datetime.date():
                    if isinstance(event_start, datetime.datetime) and isinstance(event_end, datetime.datetime):
                        print(f"[DEBUG] {start_datetime.strftime('%d.%m.%Y')}, {event_start.strftime('%H:%M')} bis {event_end.strftime('%H:%M')} '{event_summary}' wird gelöscht, weil ungleich '{full_title}'.")
                    else:
                        print(f"[DEBUG] {start_datetime.strftime('%d.%m.%Y')}, '{event_summary}' wird gelöscht.")
                    event.delete()

            # If the event does not exist, create it with all the information collected
            if not event_exists:
                # Create the description by including the break time (if available) and the task
                description = f"Dienst: {title}, "
                if workplace:
                    description += f"Platz: {workplace}, "
                if break_time:
                    description += f"Pause: {break_time}, "
                if task:
                    description += f"Aufgabe: {task}. "
                if not workplace:
                    print(f"[DEBUG] Keinen Platz für '{title}' am {start_datetime.strftime('%a, %d.%m.%Y')}")

                # Add the current modification date
                last_modified = datetime.datetime.now().strftime('%d.%m.%Y, %H:%M')
                description += "Alle Angaben und Inhalte sind ohne Gewähr. "
                description += f"Änderungsdatum: {last_modified}"

                # Create the iCal event with the full description
                # if is_holiday_flag:
                #    print(f"[DEBUG] {start_date.strftime('%a, %d.%m.%Y')} ist ein Feiertag oder Wochenende: {holiday_name}")
                ical_data = create_ical_event(
                    full_title, start_datetime, end_datetime, description=description
                )
                if ical_data:
                    cal_cache.add_event(ical_data)
                    print(f"[Dienst] {start_datetime.strftime('%d.%m.%Y')}, "
                          f"{start_datetime.strftime('%H:%M')} bis {end_datetime.strftime('%H:%M')}: {full_title}")
        except Exception as e:
            print(f"[ERROR] Fehler beim Speichern oder Löschen des Events: {full_title}, {e}")


def process_excel_file(file_path, user_name, laufzettel_werktags, laufzettel_we, countnightshifts):
    global nextlaufzettel, current_laufzettel
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active # Nimmt das erste/aktive Blatt
    rows = list(ws.iter_rows(values_only=True))

    identifier_row = None
    user_row = None

    def clean_excel_name(n):
        """Bereinigt den Namen in der Excel-Tabelle."""
        if not n: return ""
        try:
            if isinstance(n, bytes):
                n = n.decode('utf-8', errors='ignore')
            n = str(n)
        except Exception:
            return ""
            
        # Entferne Inhalte in Klammern (z.B. "(TV)", "(fester Freier)")
        n = re.sub(r'\s*\(.*\)', '', n)
        # Ersetze geschützte Leerzeichen (\xa0), Zero-Width Spaces (\u200b) und BOM
        n = n.replace('\u00A0', ' ').replace('\u200b', '').replace('\ufeff', '')
        # Reduziere mehrere Leerzeichen auf eines und trimme
        n = re.sub(r'\s+', ' ', n)
        return n.strip().lower()

    user_name_cleaned = re.sub(r',\s*[A-Z]\.?$', '', user_name).strip()
    target_full = clean_excel_name(user_name)
    target_short = clean_excel_name(user_name_cleaned)
    # print(f"[DEBUG] {user_name}, {target_full}, {target_short}")

    # 1. Finde die Zeile mit den Datumswerten (enthält "I" in Spalte A/Index 0)
    for row in rows:
        # Wir suchen nach dem römischen I oder ähnlichen Markern in der ersten Spalte
        if row[0] and isinstance(row[0], str) and "I" in row[0]:
            identifier_row = row
            break
            
    if identifier_row is None:
        print(f"[ERROR] Konnte Identifikationszeile (Datumszeile) in {os.path.basename(file_path)} nicht finden.")
        return nextlaufzettel, current_laufzettel

    for row in rows:
        cell_val = row[0]
        if cell_val:
            row_name_clean = clean_excel_name(cell_val)
            
            if not row_name_clean: 
                continue

            # A) Exakter Match (bereinigt)
            if row_name_clean == target_full:
                user_row = row
                # print(f"[DEBUG] User in Excel gefunden: {cell_val}")
                break
            
            if row_name_clean == target_short:
                user_row = row
                print(f"[DEBUG] Kurzer User in Excel gefunden: {cell_val}")
                break

    # 3. Wenn User NICHT gefunden wurde -> Lösche Logik (Urlaub/Krank/Nicht eingeteilt)
    if user_row is None:
        if not dienste:
            print(f"[WARNING] Benutzer '{user_name}' (Clean: '{target_full}') in Datei '{os.path.basename(file_path)}' nicht gefunden.")
            print(f"[DEBUG] Skript geht davon aus, dass in dieser Woche keine Dienste stattfinden -> Lösche vorhandene Termine.")
        
        # Spalten B bis H entsprechen Index 1 bis 7 in der Row-Liste
        for day_idx in range(1, 8):
            try:
                date_val = identifier_row[day_idx]
                if date_val is None: continue
                
                start_datetime = to_python_datetime(date_val)

                # Prüfe Laufzettelwechsel auch hier, um korrekten Kontext zu haben
                if nextlaufzettel and isinstance(nextlaufzettel, datetime.datetime) and start_datetime.date() >= nextlaufzettel.date():
                   if current_laufzettel:
                        old_html_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 
                                                        'Laufzettel_' + current_laufzettel.strftime('%Y%m%d') + '.html')
                        if old_html_file_path in laufzettel_cache:
                            del laufzettel_cache[old_html_file_path]
                   html_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 
                                                'Laufzettel_' + nextlaufzettel.strftime('%Y%m%d') + '.html')
                   laufzettel_werktags, laufzettel_we = parse_html_for_workplace_info_with_cache(html_file_path)
                   current_laufzettel = nextlaufzettel
                   getnextlaufzettel()

                # Suche nach existierenden Terminen an diesem Tag um sie zu löschen
                existing_events = cal_cache.get_events_on_date(start_datetime.date())
                
                for event in existing_events[:]: # Kopie der Liste iterieren
                    event_summary = event.vobject_instance.vevent.summary.value
                    event_start = event.vobject_instance.vevent.dtstart.value
                    
                    if isinstance(event_start, datetime.datetime):
                        event_start_date = event_start.date()
                    else:
                        event_start_date = event_start
                        
                    if event_start_date == start_datetime.date():
                        print(f"[DEBUG] Lösche '{event_summary}' vom {event_start_date.strftime('%d.%m.%Y')}, da Nutzer nicht in Excel gefunden.")
                        cal_cache.delete_event(event)
            except Exception as e:
                print(f"[ERROR] Fehler beim Löschen der Termine: {e}")
        return nextlaufzettel, current_laufzettel

    # 4. User gefunden -> Verarbeite Zeile
    first_date = to_python_datetime(identifier_row[1]) 
    year = first_date.year
    start_of_january = datetime.datetime(year, 1, 1, 0, 0, tzinfo=tz_berlin)
    end_of_march = datetime.datetime(year, 3, 31, 23, 59, tzinfo=tz_berlin)

    # Check für Nachtschichten-Reset (Januar Logik)
    try:
        # Kurze Prüfung, ob schon Termine im Januar existieren (um Zähler zu resetten oder nicht)
        # Das ist eine Näherung.
        termine = [] 
        # Wir nutzen hier den Cache nicht effizient für eine Jahresabfrage, 
        # aber da es nur um das Vorhandensein geht, ist es ok.
        # Alternativ: Prüfen ob events_by_date Einträge im Januar hat.
        for d, evts in cal_cache.events_by_date.items():
            if d.year == year and d.month <= 3:
                 termine.extend(evts)
        
        if len(termine) == 0:
            nonightshifts = True
        else:
            nonightshifts = False
    except Exception as e:
        print(f"[ERROR] Fehler beim Nachtschicht-Check: {e}")
        nonightshifts = True

    latest_date = None
    
    # Iteriere über Spalten Index 1 bis 7 (Montag bis Sonntag)
    for day_idx in range(1, 8):
        date_val = identifier_row[day_idx]
        service_entry = user_row[day_idx]
        
        if date_val is None:
            continue
            
        start_date = to_python_datetime(date_val)
        
        if latest_date is None or start_date.date() > latest_date:
            latest_date = start_date.date()
            # Laufzettelwechsel Check
            if nextlaufzettel and latest_date >= nextlaufzettel.date():
                if current_laufzettel:
                     old_html_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 
                                                     'Laufzettel_' + current_laufzettel.strftime('%Y%m%d') + '.html')
                     if old_html_file_path in laufzettel_cache:
                         del laufzettel_cache[old_html_file_path]
                html_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 
                                            'Laufzettel_' + nextlaufzettel.strftime('%Y%m%d') + '.html')
                laufzettel_werktags, laufzettel_we = parse_html_for_workplace_info_with_cache(html_file_path)
                current_laufzettel = nextlaufzettel
                getnextlaufzettel()

        # Eintrag bereinigen
        if service_entry is None: 
            service_entry = "FT"
        elif isinstance(service_entry, float):
             service_entry = "FT" # Fangt NaNs ab
        elif not isinstance(service_entry, str):
            service_entry = "FT"

        service_entry = service_entry.replace('\n', ' ').replace('\r', ' ')
        service_entry = re.sub(r'\s+', ' ', service_entry) # Doppel-Leerzeichen weg
        # Formate wie 10.00 zu 10:00 reparieren
        service_entry = re.sub(r'(\b\d{2})\.(\d{2}\b)', r'\1:\2', service_entry)
        service_entry = re.sub(r'(\b\d{2}:\d{2})\s*-\s*(\d{2}:\d{2}\b)', r'\1 - \2', service_entry)

        # Laufzettel Update Logik nochmal explizit für den Tag
        if nextlaufzettel is not None:
            if (start_date.date() >= nextlaufzettel.date()):
                html_file_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 
                                            'Laufzettel_' + nextlaufzettel.strftime('%Y%m%d') + '.html')
                laufzettel_werktags, laufzettel_we = parse_html_for_workplace_info_with_cache(html_file_path)
                getnextlaufzettel()
        # print(f"[DEBUG] Excel event: {service_entry}, start: {start_date}")
        # Unterscheidung: Uhrzeit (09:00 - 17:00) vs Ganztag (FT, Urlaub, etc.)
        if re.search(r'\b\d{2}:\d{2}\s*-\s*\d{2}:\d{2}\b', service_entry):
            process_timed_event(
                service_entry, start_date,
                laufzettel_werktags, laufzettel_we,
                countnightshifts, nonightshifts
            )
        else:
            process_all_day_event(service_entry, start_date)
            
    return nextlaufzettel, current_laufzettel

def load_all_laufzettel(folder_path):
    global laufzettel_data  # Zugriff auf die globale Variable
    html_files = [f for f in os.listdir(folder_path) if re.match(r'Laufzettel_\d{8}\.html', f)]
    html_files.sort()
    laufzettel_data = {}
    for html_file in html_files:
        file_path = os.path.join(folder_path, html_file)
        with open(file_path, 'r', encoding='utf-8') as file:
            laufzettel_data[html_file] = file.read()
    return


def initialize_laufzettel():
    global nextlaufzettel, current_laufzettel, laufzettel_data  # Zugriff auf die globalen Variablen
    today = date.today()
    
    # print(f"[DEBUG] Mehrere HTML-Dateien gefunden: {html_files}")
    # Sammle alle Laufzettel-Daten
    laufzettel_dates = []
    for html_file in laufzettel_data:
        try:
            laufzettel_datum = datetime.datetime.strptime(
                re.search(r'Laufzettel_(\d{8})\.html', html_file).group(1),
                "%Y%m%d"
            )
            laufzettel_dates.append(laufzettel_datum)
        except Exception as e:
            print(f"[ERROR] Fehler beim Parsen des Datums aus {html_file}: {e}")
            continue
    
    # Finde den aktuellen Laufzettel, indem du das erste Laufzettel-Datum suchst
    valid_current = [d for d in laufzettel_dates if d.date() <= today]
    if valid_current:
        current_laufzettel = min(valid_current)
        
        # Finde den nächsten Laufzettel (erster nach dem aktuellen)
        valid_next = [d for d in laufzettel_dates if d.date() > current_laufzettel.date()]
        if valid_next:
            nextlaufzettel = min(valid_next)
    
    if current_laufzettel:
        html_file_path = os.path.join(folder_path, f'Laufzettel_{current_laufzettel.strftime("%Y%m%d")}.html')
        # print(f"[DEBUG] Erster Laufzettel: {current_laufzettel.strftime('%d.%m.%Y')}")
        # if nextlaufzettel:
        #     print(f"[DEBUG] Nächster Laufzettel ab: {nextlaufzettel.strftime('%d.%m.%Y')}")
        laufzettel_werktags, laufzettel_we = parse_html_for_workplace_info_with_cache(html_file_path)
        return laufzettel_werktags, laufzettel_we
    return None, None


def getnextlaufzettel():
    global nextlaufzettel, laufzettel_data  # Zugriff auf die globale Variable
    """Bestimmt den chronologisch nächsten verfügbaren Laufzettel."""
    if nextlaufzettel is None:
        print("[WARNING] Eingabe-Laufzettel ist None")
        return None

    # Sammle alle Laufzettel-Daten
    laufzettel_dates = []
    for html_file in laufzettel_data:
        try:
            laufzettel_datum = datetime.datetime.strptime(
                re.search(r'Laufzettel_(\d{8})\.html', html_file).group(1), 
                "%Y%m%d"
            )
            laufzettel_dates.append(laufzettel_datum)
        except Exception as e:
            print(f"[ERROR] Fehler beim Parsen des Datums aus {html_file}: {e}")
            continue

    # Sortiere die Daten chronologisch
    laufzettel_dates.sort()
    
    # Finde den nächsten Laufzettel nach dem aktuellen
    valid_next = [d for d in laufzettel_dates if d.date() > nextlaufzettel.date()]
    if valid_next:
        nextlaufzettel = min(valid_next)
        # print(f"[DEBUG] Nächster Laufzettel gefunden: {nextlaufzettel.strftime('%d.%m.%Y')}")
        return nextlaufzettel
    
    # print(f"[DEBUG] Kein weiterer Laufzettel nach {nextlaufzettel.strftime('%d.%m.%Y')} gefunden")
    return None


def extract_date(entry):
    # Suche nach dem Datum im Format TT.MM.JJJJ
    match = re.search(r'(\d{2}\.\d{2}\.\d{4})', entry)
    if match:
        date_str = match.group(1)
        # Datum in das Format JJJJMMTT umwandeln, um lexikografische Sortierung zu ermöglichen
        return date_str[6:] + date_str[3:5] + date_str[:2]  # Format: JJJJMMTT
    return None


def send_email(subject, body, to_email, kalender_id):
    from_email = load_credentials("notifymail", config_path)
    from_display_name = "Dein Dienstplan"
    password = load_credentials("mailpassword", config_path)
    kalenderbase = load_credentials("kalenderbase", config_path)
    abobase = load_credentials("abobase", config_path)
    night_shifts = None
    if date.today().month >= 11:
        night_shifts_count = count_night_shifts(date.today())
        night_shifts_count_year = count_night_shifts(datetime.datetime(date.today().year, 12, 31, 23, 59))
        if night_shifts_count > 0 or night_shifts_count_year > 0:
            if night_shifts_count == 1:
                night_shifts = f"Im Jahr {date.today().year} hattest du bisher {night_shifts_count} Nachtschicht."
            elif night_shifts_count > 1:
                night_shifts = f"Im Jahr {date.today().year} hattest du bisher {night_shifts_count} Nachtschichten."
            else:
                night_shifts = f"Im Jahr {date.today().year} hattest du bisher keine Nachtschichten."
            if night_shifts_count_year - night_shifts_count > 0:
                if night_shifts_count_year - night_shifts_count == 1:
                    night_shifts += f"<br>Es sind noch {night_shifts_count_year - night_shifts_count} Nachtschicht für dich disponiert.<br>"
                elif night_shifts_count_year - night_shifts_count > 1:
                    night_shifts += f"<br>Es sind noch {night_shifts_count_year - night_shifts_count} Nachtschichten für dich disponiert.<br>"
                if night_shifts_count_year == 1:
                    night_shifts += f"Das ist dann insgesamt {night_shifts_count_year} Nachtschicht für {date.today().year}."
                else:
                    night_shifts += f"Das wären dann insgesamt {night_shifts_count_year} Nachtschichten für {date.today().year}."
    kalenderurls = (
        f'<a href="{kalenderbase}{kalender_id}">Kalender</a><br>'
        f'<a href="{abobase}{kalender_id}?export">Abo-URL</a><br>Alle Angaben und Inhalte sind ohne Gewähr.'
    )
    if night_shifts:
        body += "<br><br>" + night_shifts
    body += "<br><br>" + kalenderurls
    msg = MIMEMultipart()
    msg['From'] = f"{from_display_name} <{from_email}>"
    msg['To'] = to_email
    msg['Subject'] = subject
    msg.attach(MIMEText(body, 'html'))
    try:
        server = smtplib.SMTP('smtp.ionos.de', 587)
        server.starttls()
        server.login(from_email, password)
        text = msg.as_string()
        server.sendmail(from_email, to_email, text)
        server.quit()
        print(f"[INFO] E-Mail an {to_email} gesendet.")
    except Exception as e:
        print(f"[ERROR] Fehler beim Senden der E-Mail: {e}")


def load_email_config(user_name, email_config_path):  # Funktion zum Laden der E-Mail-Konfiguration
    try:
        with open(email_config_path, 'r') as file:
            email_config = json.load(file)

        if user_name in email_config:
            return email_config[user_name]
        else:
            raise ValueError(f"Keine E-Mail-Konfiguration für '{user_name}' gefunden.")
    except FileNotFoundError:
        raise FileNotFoundError(f"Die Datei '{email_config_path}' wurde nicht gefunden.")
    except Exception as e:
        raise ValueError(f"Fehler beim Laden der E-Mail-Konfiguration: {e}")


def load_credentials(service_name, config_path):
    # Laden der JSON-Datei
    with open(config_path, 'r') as file:
        config = json.load(file)

    # Der Schlüssel in der config.json entspricht direkt dem service_name (z.B. "username_mm" oder "password_mm")
    if not "_" in service_name:
        if service_name in config:
            return config[service_name]  # Nur das Passwort zurückgeben
        else:
            raise ValueError(f"Config für '{service_name}' nicht gefunden.")
    else:
        # Allgemeiner Fall für andere Dienste
        username_key = f"username_{service_name}"
        password_key = f"password_{service_name}"

        # Überprüfen, ob der Benutzername und das Passwort existieren
        if username_key in config and password_key in config:
            username = config[username_key]
            password = config[password_key]
            return username, password
        else:
            raise ValueError(f"Benutzername oder Passwort für '{service_name}' nicht gefunden.")


def extract_date_from_filename(filename):
    # Regulärer Ausdruck für die Datumsangaben im Dateinamen
    match = re.search(r'(\d{1,2})\.(\d{1,2})\.(\d{2,4}) - (\d{1,2})\.(\d{1,2})\.(\d{2,4})', filename)
    if match:
        start_day = int(match.group(1))
        start_month = int(match.group(2))
        start_year = int(match.group(3))
        if start_year < 100:
            start_year += 2000
        start_date = dt.datetime(start_year, start_month, start_day)  # Verwende 'dt.datetime'
        return start_date
    else:
        return None  # Falls das Datum nicht gefunden wird


# Main
locale.setlocale(locale.LC_TIME, 'de_DE.UTF-8')
tz_berlin = pytz.timezone('Europe/Berlin')
script_path = os.path.abspath(__file__)
folder_path = os.path.dirname(script_path)
# print(f"[DEBUG] folder_path: {folder_path}")
email_config_path = os.path.join(folder_path, 'email_config.json')
config_path = os.path.join(folder_path, 'config.json')
# print(f"[DEBUG] Config: {config_path}")
night_shifts_count = 0
countnightshifts = False  # Initialisierung vor der Verarbeitung
changedlaufzettel = False
laufzettel_cache = {}
laufzettel_data = {}
# Globale Variablen für Laufzettel
nextlaufzettel = None
current_laufzettel = None

parser = argparse.ArgumentParser(description="Dienst zu ARD ZDF Box Script.")
parser.add_argument("user_name", help="Name des Benutzers", type=str)
parser.add_argument("-c", "--calendar", help="Dienste in MeiersMarkus eintragen", action="store_true")
parser.add_argument("-o", "--ort", help="Ort als Adresse hinzufügen", action="store_true")
parser.add_argument("-r", "--rewrite", help="Alle vorhandenen Termine im Zeitbereich neu erstellen", action="store_true")
parser.add_argument("-n", "--notify", help="E-Mail mit eingetragenen Diensten an Benutzer senden", action="store_true")
parser.add_argument("-s", "--nas", help="Direkt aufs NAS", action="store_true")
parser.add_argument("-d", "--dienste", help="Nur Dienste eintragen", action="store_true")

args = parser.parse_args()
user_name = args.user_name
meiersmarkus = args.calendar
ort = args.ort
rewrite = args.rewrite
notify = args.notify
nas = args.nas
dienste = args.dienste

if meiersmarkus:
    service_name = "mm"
elif nas:
    service_name = "nas"
else:
    service_name = "ard"

caldavlogin = "caldav" + service_name
caldav_start = load_credentials(caldavlogin, config_path)
calendar_name = 'Dienstplan ' + user_name.replace(',', '').replace('.', '')
# 2. URL Generierung (Umlaute ersetzen: ü -> ue, ß -> ss)
# Bereinige den Namen für die URL: Kleinbuchstaben, Umlaute weg, Leerzeichen zu Strichen
# clean_user_name = replace_german_umlauts(user_name).lower()
clean_user_name = user_name.replace(' ', '-').replace(',', '').replace('.', '')
clean_user_name = encode_calendar_url(clean_user_name)

caldav_url = caldav_start + 'dienstplan-' + clean_user_name + '/'

# print(f"[DEBUG] Kalendername: {calendar_name}")
# print(f"[DEBUG] Kalender-URL: {caldav_url}")

options = [
    "'-c'" if meiersmarkus else None,
    "'-o'" if ort else None,
    "'-r'" if rewrite else None,
    "'-n'" if notify else None,
    "'-s'" if nas else None,
    "'-d'" if dienste else None
]
options = [opt for opt in options if opt]
if options:
    print(f"[DEBUG] Benutzername: {user_name}, " + ", ".join(options) + "")
else:
    print(f"[DEBUG] Benutzername: {user_name}")

current_year = date.today().year
years = [current_year - 1, current_year, current_year + 1]
de_holidays = holidays.Germany(years=years, observed=False, prov="HH", language="de")
for year in years:
    de_holidays[date(year, 10, 31)] = "Reformationstag"  # Reformationstag
    de_holidays[date(year, 12, 24)] = "Heiligabend"  # Heiligabend
    de_holidays[date(year, 12, 31)] = "Silvester"    # Silvester
    de_holidays[easter(year)] = "Ostersonntag"       # Ostersonntag
    de_holidays[easter(year) + timedelta(days=49)] = "Pfingstsonntag"  # Pfingstsonntag
# for holiday_date, holiday_name in de_holidays.items():
#    print(f"{holiday_date.strftime('%d.%m.%Y')}: {holiday_name}")

start_timer("caldav")

# 1. Bereinige den Namen für den Vergleich (nicht für die URL!)
target_display_name = 'Dienstplan ' + user_name.replace(',', '').replace('.', '').strip()
target_display_name_clean = " ".join(target_display_name.split()).lower()

# print(f"[DEBUG] Suche Kalender mit dem Namen: '{target_display_name}'")

try:
    # Login Daten laden
    login_service = "login_" + service_name
    username, password = load_credentials(login_service, config_path)
    
    # Verbindung zum Account herstellen (nicht zu einem spezifischen Kalender!)
    client = DAVClient(caldav_start, username=username, password=password)
    principal = client.principal()
    
    # Alle Kalender abrufen
    calendars = principal.calendars()
    
    calendar = None
    
    # Suchen...
    for cal in calendars:
        if not cal.name: 
            continue
            
        # Namen vom Server holen und bereinigen
        server_cal_name = str(cal.name)
        server_cal_name_clean = " ".join(server_cal_name.split()).lower()
        
        # Vergleich 1: Exakter Name
        if server_cal_name_clean == target_display_name_clean:
            calendar = cal
            # print(f"[SUCCESS] Kalender gefunden: '{server_cal_name}' (URL: {cal.url})")
            break
        
        # Vergleich 2: Fallback für "Preuß" vs "Preuss" (Encoding Toleranz)
        # Falls Python "ß" anders sieht als der Server
        if replace_german_umlauts(server_cal_name_clean) == replace_german_umlauts(target_display_name_clean):
            calendar = cal
            # print(f"[SUCCESS] Kalender (via Umlaut-Match) gefunden: '{server_cal_name}'")
            break

    # Wenn immer noch nicht gefunden und wir NICHT im NAS Modus sind -> Neu anlegen
    if not calendar and not nas:
        print(f"[INFO] Kalender '{target_display_name}' existiert nicht.")
        # try:
            # calendar = principal.make_calendar(name=target_display_name)
            # print(f"[SUCCESS] Neuer Kalender erstellt: '{target_display_name}'")
        # except Exception as create_err:
            # print(f"[ERROR] Konnte Kalender nicht erstellen: {create_err}")

except Exception as e:
    print(f"[ERROR] Genereller Fehler bei der CalDAV-Verbindung: {e}")
    if not nas:
        sys.exit(1)

# Letzte Prüfung
if not calendar:
    print(f"[ERROR] Kalender '{target_display_name}' konnte endgültig nicht gefunden werden.")
    sys.exit(1)

end_timer("caldav", "Verbindung zu CalDAV")

# Initialisiere Cache
cal_cache = CalendarCache(client, calendar)

# Definiere Zeitraum: 1. Januar aktuelles Jahr bis Ende nächsten Jahres (zur Sicherheit)
cache_start = datetime.datetime(current_year, 1, 1, 0, 0)
cache_end = datetime.datetime.now() + datetime.timedelta(days=90)

# Cache einmalig füllen
cal_cache.load_all_events(cache_start, cache_end)

eingetragene_termine = []
target_folder = os.path.join(folder_path, "Plaene", "MAZ_TAZ Dienstplan")
load_all_laufzettel(folder_path)
laufzettel_werktags, laufzettel_we = initialize_laufzettel()

end_timer("initial", "Initialisierung")
# print(f"[DEBUG] Absoluter Pfad zum Skript: {script_path}")
# print(f"[DEBUG] Absoluter Pfad zum Zielordner: {target_folder}")

xlsx_files = [
    os.path.join(root, f)
    for root, _, files in os.walk(target_folder)
    for f in files if f.endswith('.xlsx')
]
with_date, without_date = [], []
for file in xlsx_files:
    date = extract_date_from_filename(os.path.basename(file))
    (with_date if date else without_date).append((file, date) if date else file)
with_date.sort(key=lambda x: x[1])
xlsx_files = list(chain((f[0] for f in with_date), without_date))

if xlsx_files:
    # print(f"[DEBUG] Anzahl der gefundenen .xlsx-Dateien in {root}: {len(xlsx_files)}")
    # print(f"[DEBUG] Anzahl der .xlsx-Dateien: {len(xlsx_files)}")
    for file_path in xlsx_files:
        file_name = os.path.basename(file_path)
        # start_timer("xlsx")
        # print(f"[DEBUG] Verarbeite Datei: {file_name}")
        process_excel_file(
            file_path, user_name, laufzettel_werktags, laufzettel_we, countnightshifts
        )
        # end_timer("xlsx", f"Verarbeitung der Excel-Datei {file_name}")
else:
    print("[DEBUG] Keine .xlsx-Dateien gefunden.")
# eingetragene_termine.sort(key=lambda x: extract_date(x))
if eingetragene_termine and notify:
    print(f"[INFO] {len(eingetragene_termine)} neue Termine eingetragen.")
    eingetragene_termine_wochentag = []
    deutsche_wochentage = ['Mo.', 'Di.', 'Mi.', 'Do.', 'Fr.', 'Sa.', 'So.']
    for term in eingetragene_termine:
        split_term = term.split(":")
        split_term = split_term[0].split(" ")
        try:
            datum = datetime.datetime.strptime(split_term[0], '%d.%m.%Y')
        except ValueError:
            continue # Überspringen bei Formatfehler
        wochentag = deutsche_wochentage[datum.weekday()] # 0=Montag, 6=Sonntag
        # wochentag = pd.to_datetime(split_term[0], format='%d.%m.%Y').strftime('%a')[:2] + '.'
        # print(f"[INFO] {wochentag} {term}")
        # Wochentag vor jedes Datum einfügen und in die Liste eingetragene_termine_wochentag schreiben
        # Wenn das Datum in der Vergangenheit liegt, dann nicht einfügen
        if datum.date() >= datetime.date.today():
            eingetragene_termine_wochentag += [f"{wochentag} {term}"]
    # start_timer("mail")
    # Zusammenstellung der eingetragenen Termine
    if eingetragene_termine_wochentag:
        mail_body = "Es wurden folgende Termine eingetragen:<br><br>"
        mail_body += "<br>".join(eingetragene_termine_wochentag)
        # Laden der E-Mail-Konfiguration
        try:
            email_config = load_email_config(user_name, email_config_path)
            send_email(
                subject=f"Dienstplan Update {user_name}",
                body=mail_body,
                to_email=email_config["email"],
                kalender_id=email_config["kalender_id"]
            )
            print(f"[INFO] Mail über {len(eingetragene_termine_wochentag)} neue Termine abgeschickt.")
        except Exception as e:
            print(f"[ERROR] Fehler bei der Benachrichtigung: {e}")
    # end_timer("mail", "Mail")
elif eingetragene_termine and not notify:
    print(f"[INFO] {len(eingetragene_termine)} neue Termine eingetragen.")
# else:
#    print("[INFO] Keine neuen Termine eingetragen.")
end_timer("gesamt", "Zeit")
